"""
대법원 회생사건 자동 조회 크롤러 v11.2
- 자동입력방지문자(보안문자) 자동 해제
- 진행내용 전체 파싱 + 글자색 반영 + 분류 컬럼(v11.1)
- 플래그 자동 분류 (v11: 13개 -> 17개 항목으로 확대)
- 조회일시/결과 5행 표시
- 맨 앞 "회생차주 총괄표" 시트 자동 생성 (주요 일정 + 진행상황)
- 총괄표 진행상황 산정 기준 변경: 종국결과 > 인가결정 > 개시결정 > 개시신청
- 파산(하합/하단)·개인회생(개회) 사건번호 지원
설치: pip install selenium openpyxl 2captcha-python webdriver-manager python-dotenv pandas
.env_verify: VERIFY_API_KEY=실제키입력
"""

import time
import re
import os
from datetime import datetime
from dotenv import load_dotenv
from selenium import webdriver
from selenium.common.exceptions import NoAlertPresentException, TimeoutException
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import Select, WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.chrome.service import Service
from webdriver_manager.chrome import ChromeDriverManager
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import pandas as pd
from twocaptcha import TwoCaptcha as _VerifySolver

load_dotenv(dotenv_path=".env_verify")
VERIFY_API_KEY = os.getenv("VERIFY_API_KEY", "") or os.getenv("TWOCAPTCHA_API_KEY", "")
if not VERIFY_API_KEY:
    try:
        from dotenv import dotenv_values
        _vals = {**dotenv_values(".env_verify"), **dotenv_values(".env_capcha")}
        VERIFY_API_KEY = _vals.get("VERIFY_API_KEY", "") or _vals.get("TWOCAPTCHA_API_KEY", "")
    except Exception:
        pass

BTPR_NM = "아이"

INPUT_EXCEL_PATH  = r"C:\Users\LEE CPA\OneDrive\Desktop\Record\python Project\5. 나의사건검색 크롤링\INPUT DATA.xlsx"
INPUT_SHEET_NAME  = "Sheet1"
COL_SERIAL        = "A"
COL_NAME          = "B"
COL_COURT         = "C"
COL_CASE_NUMBER   = "D"
HEADER_ROW        = 1
OUTPUT_EXCEL_PATH = f"회생사건_조회결과_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
DELAY_BETWEEN_REQUESTS = 1.0
MAX_CAPTCHA_RETRIES    = 5

SUMMARY_SHEET_NAME = "회생차주 총괄표"

TARGET_URL     = "https://ssgo.scourt.go.kr/ssgo/index.on?cortId=www"
ID_COURT       = "mf_ssgoTopMainTab_contents_content1_body_sbx_cortCd"
ID_FULL_CS_NO  = "mf_ssgoTopMainTab_contents_content1_body_ibx_fullCsNo"
ID_BTPR_NM     = "mf_ssgoTopMainTab_contents_content1_body_ibx_btprNm"
ID_SANO_MODE   = "mf_ssgoTopMainTab_contents_content1_body_cbx_chkSanoInputMode_input_0"
ID_CAPTCHA_IMG = "mf_ssgoTopMainTab_contents_content1_body_img_captcha"
ID_CAPTCHA_INP = "mf_ssgoTopMainTab_contents_content1_body_ibx_answer"
ID_SEARCH_BTN  = "mf_ssgoTopMainTab_contents_content1_body_btn_srchCs"
ID_TAB2        = "mf_ssgoTopMainTab_contents_content1_body_wfSsgoDetail_ssgoCsDetailTab_tab_ssgoTab2_tabHTML"

FLAG_COLS = [
    "신청서접수", "답변서제출", "심문사항답변", "심문기일", "현장검증기일",
    "급여", "시부인표", "채권신고및변제요청서", "월간보고서", "기타보고서",
    "연간보고서", "감사의견서",
    "조사보고서", "회생계획안", "임금채권공익채권",
    "세금납부허가", "신탁담보변경",
]

FLAG_HEADER_COLS = ["분류", "당부사항", "열람대상"] + FLAG_COLS

COLOR_CATEGORY = {
    "0000CC66": "송달", "00CC6600": "송달",
    "00660000": "제출서류",
    "00336633": "명령",
    "00000000": "공고",
    "00003399": "기일", "000033FF": "기일",
}


def categorize_color(rgb_str):
    hexc = rgb_to_hex(rgb_str)
    table = {
        "CC6600": "송달", "660000": "제출서류",
        "336633": "명령", "000000": "공고", "003399": "기일",
    }
    if hexc in table:
        return table[hexc]
    try:
        r = int(hexc[0:2], 16); g = int(hexc[2:4], 16); b = int(hexc[4:6], 16)
    except Exception:
        return ""
    if r < 40 and g < 40 and b < 40:
        return "공고"
    if b > 120 and b > r and b > g:
        return "기일"
    if r > 150 and 60 <= g <= 150 and b < 60:
        return "송달"
    if r > 80 and g < 60 and b < 60:
        return "제출서류"
    if g > 80 and r < 90 and b < 90:
        return "명령"
    return ""


def setup_driver():
    options = webdriver.ChromeOptions()
    options.add_argument("--no-sandbox")
    options.add_argument("--disable-dev-shm-usage")
    options.add_argument("--window-size=1280,900")
    options.add_argument("user-agent=Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36")
    driver = webdriver.Chrome(service=Service(ChromeDriverManager().install()), options=options)
    driver.set_page_load_timeout(30)
    return driver


_HEADER_KEYWORDS = ["일련번호", "차주일련번호", "차주명", "관할법원", "회생사건번호", "사건번호"]
_CASE_HINT = re.compile(r'\d{4}(간회합|간회단|회합|회단|간회|개회|하합|하단|간이|하)\d+')


def _looks_like_header(cells):
    joined = "".join(cells).replace(" ", "")
    if _CASE_HINT.search(joined):
        return False
    keyset = [k.replace(" ", "") for k in _HEADER_KEYWORDS]
    hits = sum(1 for c in cells if c.replace(" ", "") in keyset)
    return hits >= 2


def parse_pasted_table(text):
    cases = []
    lines = text.replace("\r\n", "\n").replace("\r", "\n").split("\n")
    for ln in lines:
        if not ln.strip():
            continue
        if "\t" in ln:
            cells = [c.strip() for c in ln.split("\t")]
        elif "," in ln:
            cells = [c.strip() for c in ln.split(",")]
        else:
            cells = [c.strip() for c in re.split(r"\s{2,}", ln)]
        if len(cells) < 4:
            continue
        if _looks_like_header(cells):
            continue
        serial, name, court, case = cells[0], cells[1], cells[2], cells[3]
        if not case or case.lower() == "none":
            continue
        cases.append({"일련번호": serial, "차주명": name, "법원": court, "사건번호": case})
    return cases


def load_cases(excel_path, sheet_name, col_serial, col_name, col_court, col_case, header_row):
    wb = openpyxl.load_workbook(excel_path, data_only=True)
    ws = wb[sheet_name]
    ci_serial = openpyxl.utils.column_index_from_string(col_serial) - 1
    ci_name   = openpyxl.utils.column_index_from_string(col_name)   - 1
    ci_court  = openpyxl.utils.column_index_from_string(col_court)  - 1
    ci_case   = openpyxl.utils.column_index_from_string(col_case)   - 1
    cases = []
    for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
        serial = str(row[ci_serial]).strip() if row[ci_serial] else ""
        name   = str(row[ci_name]).strip()   if row[ci_name]   else ""
        court  = str(row[ci_court]).strip()  if row[ci_court]  else ""
        case   = str(row[ci_case]).strip()   if row[ci_case]   else ""
        if case and case != "None":
            cases.append({"일련번호": serial, "차주명": name, "법원": court, "사건번호": case})
    print(f"총 {len(cases)}건 로드 완료")
    return cases


CASE_TYPE_RE = re.compile(
    r'(\d{4})'
    r'(간회합|간회단|회합|회단'
    r'|간회|개회'
    r'|하합|하단|간이|하'
    r')'
    r'(\d+)$'
)


def parse_case_number(s):
    m = CASE_TYPE_RE.search(s.strip())
    if m:
        return {"year": m.group(1), "type_code": m.group(2), "serial": m.group(3)}
    return None


def rgb_to_hex(rgb_str):
    try:
        nums = re.findall(r'\d+', rgb_str)
        return '{:02X}{:02X}{:02X}'.format(int(nums[0]), int(nums[1]), int(nums[2]))
    except:
        return '000000'


def wait_el(driver, el_id, timeout=10):
    return WebDriverWait(driver, timeout).until(EC.presence_of_element_located((By.ID, el_id)))


def quick_accept_alert(driver, timeout=0.4):
    end = time.time() + timeout
    while time.time() < end:
        try:
            alert = driver.switch_to.alert
            txt = alert.text
            alert.accept()
            return txt
        except NoAlertPresentException:
            time.sleep(0.05)
    return None


def dismiss_popup(driver, timeout=0.4):
    end = time.time() + timeout
    while time.time() < end:
        try:
            btns = driver.find_elements(By.XPATH, "//button[contains(text(),'확인')]")
            for btn in btns:
                if btn.is_displayed() and btn.is_enabled():
                    btn.click()
                    return True
        except:
            pass
        time.sleep(0.05)
    return False


def select_court(driver, court, timeout=15):
    try:
        sel = driver.find_element(By.ID, ID_COURT)
        if len(Select(sel).options) <= 200:
            WebDriverWait(driver, timeout).until(
                lambda d: len(Select(d.find_element(By.ID, ID_COURT)).options) > 200
            )
    except:
        WebDriverWait(driver, timeout).until(
            lambda d: len(Select(d.find_element(By.ID, ID_COURT)).options) > 200
        )
    driver.execute_script("""
        var sel = document.getElementById(arguments[0]);
        for (var i = 0; i < sel.options.length; i++) {
            if (sel.options[i].value === arguments[1] || sel.options[i].text === arguments[1]) {
                sel.selectedIndex = i;
                sel.dispatchEvent(new Event('change', {bubbles: true}));
                return;
            }
        }
    """, ID_COURT, court)


def js_set_input(driver, el_id, value):
    driver.execute_script("""
        var el = document.getElementById(arguments[0]);
        el.focus();
        var nativeInputValueSetter = Object.getOwnPropertyDescriptor(
            window.HTMLInputElement.prototype, 'value').set;
        nativeInputValueSetter.call(el, arguments[1]);
        el.dispatchEvent(new Event('input', {bubbles: true}));
        el.dispatchEvent(new Event('change', {bubbles: true}));
        el.dispatchEvent(new KeyboardEvent('keyup', {bubbles: true}));
    """, el_id, value)


def get_captcha_base64(driver):
    try:
        WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.ID, ID_CAPTCHA_IMG)))
        return driver.execute_script("""
            var img = document.getElementById(arguments[0]);
            if (!img) return null;
            var canvas = document.createElement('canvas');
            canvas.width  = img.naturalWidth  || img.width  || 130;
            canvas.height = img.naturalHeight || img.height || 50;
            var ctx = canvas.getContext('2d');
            ctx.drawImage(img, 0, 0);
            return canvas.toDataURL('image/png').split(',')[1];
        """, ID_CAPTCHA_IMG)
    except Exception as e:
        print(f"  보안문자 이미지 추출 실패: {e}")
        return None


def resolve_verification(api_key, image_b64):
    solver = _VerifySolver(api_key)
    try:
        result = solver.normal(image_b64)
        return result["code"]
    except Exception as e:
        print(f"  인증 처리 실패: {e}")
        return None


def parse_result(driver, court, case_str):
    try:
        alert_text = quick_accept_alert(driver, timeout=0.6)
        if alert_text and ("자동입력" in alert_text or "방지문자" in alert_text):
            return {"_retry": True}

        src = driver.page_source

        if any(x in src for x in ["자동입력방지문자가 올바르지", "보안문자를 다시", "captchaError"]):
            return {"_retry": True}

        if any(x in src for x in ["검색된 사건이 없습니다", "해당 사건이 없습니다", "조회된 내역이 없"]):
            return {"사건번호": case_str, "관할법원": court, "결과": "사건 없음",
                    "조회일시": datetime.now().strftime("%Y-%m-%d %H:%M"), "기본내용": {}, "진행내용": []}

        tab_ok = False
        for tab_try in range(3):
            try:
                WebDriverWait(driver, 8).until(
                    EC.presence_of_element_located((By.ID, ID_TAB2))
                )
                driver.execute_script("document.getElementById(arguments[0]).click();", ID_TAB2)
                if _wait_progress_table(driver, timeout=8.0):
                    tab_ok = True
                    break
                time.sleep(0.5)
            except Exception as e:
                if tab_try == 2:
                    print(f"  진행내용 탭 클릭 실패(3회): {e}")
                time.sleep(0.6)

        basic = {"관할법원": court}
        BASIC_FIELDS = ["사건번호", "사건명", "신청인", "채권자", "재판부", "접수일", "종국결과", "채무자"]
        try:
            for row in driver.find_elements(By.CSS_SELECTOR, "table tr"):
                ths = row.find_elements(By.TAG_NAME, "th")
                tds = row.find_elements(By.TAG_NAME, "td")
                for th, td in zip(ths, tds):
                    k = th.text.strip()
                    v = td.text.strip()
                    if k in BASIC_FIELDS and v and k not in basic:
                        basic[k] = v
        except Exception as e:
            print(f"  기본내용 파싱 실패: {e}")

        progress = []
        try:
            for table in driver.find_elements(By.CSS_SELECTOR, "table"):
                headers = [th.text.strip() for th in table.find_elements(By.TAG_NAME, "th")]
                if "일자" in headers and "내용" in headers and "결과" in headers and "공시문" in headers:
                    rows_data = driver.execute_script("""
                        var rows = arguments[0].querySelectorAll('tr');
                        var result = [];
                        for (var i = 1; i < rows.length; i++) {
                            var cells = rows[i].querySelectorAll('td');
                            if (cells.length < 2) continue;
                            var color = window.getComputedStyle(cells[1]).color;
                            result.push([
                                cells[0] ? cells[0].innerText.trim() : '',
                                cells[1] ? cells[1].innerText.trim() : '',
                                cells[2] ? cells[2].innerText.trim() : '',
                                cells[3] ? cells[3].innerText.trim() : '',
                                color
                            ]);
                        }
                        return result;
                    """, table)
                    for row_data in rows_data:
                        if row_data[0] or row_data[1]:
                            progress.append(row_data)
                    break
        except Exception as e:
            print(f"  진행내용 파싱 실패: {e}")

        if not progress:
            return {"_retry_fetch": True}

        return {"사건번호": case_str, "관할법원": court, "결과": "조회성공",
                "조회일시": datetime.now().strftime("%Y-%m-%d %H:%M"),
                "기본내용": basic, "진행내용": progress}

    except Exception as e:
        return {"사건번호": case_str, "관할법원": court, "결과": "오류", "오류내용": str(e),
                "조회일시": datetime.now().strftime("%Y-%m-%d %H:%M"), "기본내용": {}, "진행내용": []}


def _wait_progress_table(driver, timeout=8.0):
    end = time.time() + timeout
    last_count = -1
    stable = 0
    while time.time() < end:
        try:
            count = driver.execute_script("""
                var tables = document.querySelectorAll('table');
                for (var t = 0; t < tables.length; t++) {
                    var ths = tables[t].querySelectorAll('th');
                    var txt = '';
                    for (var i = 0; i < ths.length; i++) txt += ths[i].innerText;
                    if (txt.indexOf('일자') >= 0 && txt.indexOf('내용') >= 0 &&
                        txt.indexOf('결과') >= 0 && txt.indexOf('공시문') >= 0) {
                        var rows = tables[t].querySelectorAll('tr');
                        var n = 0;
                        for (var r = 1; r < rows.length; r++) {
                            if (rows[r].querySelectorAll('td').length >= 2) n++;
                        }
                        return n;
                    }
                }
                return -1;
            """)
            if count is not None and count >= 1:
                if count == last_count:
                    stable += 1
                    if stable >= 2:
                        return True
                else:
                    stable = 0
                last_count = count
        except:
            pass
        time.sleep(0.15)
    return last_count >= 1


def search_one(driver, api_key, court, case_str, btpr_nm=None):
    if btpr_nm is None:
        btpr_nm = BTPR_NM
    parsed = parse_case_number(case_str)
    if not parsed:
        return {"사건번호": case_str, "관할법원": court, "결과": "오류",
                "오류내용": "사건번호 형식 인식 불가",
                "조회일시": datetime.now().strftime("%Y-%m-%d %H:%M"), "기본내용": {}, "진행내용": []}

    for attempt in range(1, MAX_CAPTCHA_RETRIES + 1):
        try:
            driver.get(TARGET_URL)
            WebDriverWait(driver, 15).until(
                EC.presence_of_element_located((By.ID, ID_SANO_MODE))
            )
            quick_accept_alert(driver, timeout=0.2)

            sano_chk = wait_el(driver, ID_SANO_MODE)
            if not sano_chk.is_selected():
                sano_chk.click()

            select_court(driver, court)

            full_input = WebDriverWait(driver, 10).until(
                EC.element_to_be_clickable((By.ID, ID_FULL_CS_NO))
            )
            full_input.click()
            full_input.clear()
            full_input.send_keys(case_str)
            quick_accept_alert(driver, timeout=0.3)
            dismiss_popup(driver, timeout=0.2)

            js_set_input(driver, ID_BTPR_NM, btpr_nm)
            quick_accept_alert(driver, timeout=0.3)
            dismiss_popup(driver, timeout=0.2)

            print(f"  [{attempt}/{MAX_CAPTCHA_RETRIES}] 보안문자 확인 중...")
            b64 = get_captcha_base64(driver)
            if not b64:
                continue

            answer = resolve_verification(api_key, b64)
            if not answer:
                continue
            print(f"  [{attempt}/{MAX_CAPTCHA_RETRIES}] 보안문자: '{answer}'")

            js_set_input(driver, ID_CAPTCHA_INP, answer)
            dismiss_popup(driver, timeout=0.2)

            driver.execute_script("document.getElementById(arguments[0]).click();", ID_SEARCH_BTN)
            _wait_after_search(driver, timeout=4.0)

            result = parse_result(driver, court, case_str)
            if result.get("_retry"):
                print(f"  [{attempt}/{MAX_CAPTCHA_RETRIES}] 자동입력 확인 재시도...")
                continue
            if result.get("_retry_fetch"):
                print(f"  [{attempt}/{MAX_CAPTCHA_RETRIES}] 진행내용 수집 실패, 재시도...")
                time.sleep(1)
                continue

            return result

        except Exception as e:
            quick_accept_alert(driver, timeout=0.3)
            print(f"  [{attempt}/{MAX_CAPTCHA_RETRIES}] 예외: {e}")
            time.sleep(1)

    return {"사건번호": case_str, "관할법원": court, "결과": "오류",
            "오류내용": "진행내용 수집 실패 (재시도 초과)",
            "조회일시": datetime.now().strftime("%Y-%m-%d %H:%M"), "기본내용": {}, "진행내용": []}


def _wait_after_search(driver, timeout=4.0):
    end = time.time() + timeout
    while time.time() < end:
        try:
            _ = driver.switch_to.alert
            return
        except NoAlertPresentException:
            pass
        try:
            done = driver.execute_script("""
                if (document.getElementById(arguments[0])) return true;
                var b = document.body ? document.body.innerText : '';
                if (b.indexOf('검색된 사건이 없') >= 0 ||
                    b.indexOf('해당 사건이 없') >= 0 ||
                    b.indexOf('올바르지') >= 0) return true;
                return false;
            """, ID_TAB2)
            if done:
                return
        except:
            pass
        time.sleep(0.1)


def compute_flags(df: pd.DataFrame) -> pd.DataFrame:
    content_ns = df["내용"].fillna("").astype(str).str.replace(r"\s+", "", regex=True)
    result_ns  = df["결과"].fillna("").astype(str).str.replace(r"\s+", "", regex=True)

    is_ext   = content_ns.str.contains("연장", na=False)
    is_defer = content_ns.str.contains("연기", na=False)

    block = (
        ~content_ns.str.contains("송달|도달|발송|에게", na=False)
        & ~result_ns.str.contains("송달|도달", na=False)
        & ~content_ns.str.contains("기일변경|명의변경|계좌변경", na=False)
        & ~is_defer
    )

    def flag_all(*tokens, exclude=None):
        cond = ~is_ext
        for t in tokens:
            cond &= content_ns.str.contains(t, na=False)
        if exclude:
            cond &= ~content_ns.str.contains(exclude, na=False)
        return (cond & block).map(lambda x: "Y" if x else "")

    df["신청서접수"]   = flag_all("신청서접수")
    df["답변서제출"]   = flag_all("답변서", "제출", exclude="의견서")
    df["심문사항답변"] = flag_all("심문사항", "답변|보정")
    df["심문기일"] = (
        (~is_ext) & content_ns.str.contains("심문기일", na=False)
        & content_ns.str.contains("심문실|법|:", na=False)
        & ~result_ns.str.contains("변경|연기|취소", na=False)
        & block
    ).map(lambda x: "Y" if x else "")
    df["현장검증기일"] = (
        (~is_ext) & content_ns.str.contains("현장검증기일", na=False)
        & content_ns.str.contains(":", na=False)
        & ~result_ns.str.contains("변경|연기|취소", na=False)
        & block
    ).map(lambda x: "Y" if x else "")
    df["급여"] = (
        (~is_ext)
        & content_ns.str.contains("급여|임금", na=False)
        & content_ns.str.contains("지급|책정", na=False)
        & ~content_ns.str.contains("임금채권", na=False)
        & block
    ).map(lambda x: "Y" if x else "")
    df["시부인표"]             = flag_all("시부인표")
    df["채권신고및변제요청서"] = flag_all("채권신고및변제요청서")
    df["월간보고서"]           = flag_all("월간보고서")
    df["기타보고서"]           = flag_all("기타보고서", "제출")
    df["연간보고서"]           = flag_all("연간보고서")
    df["감사의견서"]           = flag_all("감사의견서")
    df["조사보고서"]           = flag_all("조사보고서")
    df["회생계획안"]           = flag_all("회생계획안",
                                  exclude="지정|연기|발송|기간|집회|기일|요지|배제")
    df["임금채권공익채권"]     = flag_all("임금채권|공익채권",
                                  exclude="신고|구매대금|리스료|렌탈료|검사료|요금|임차료")
    df["세금납부허가"] = (
        (~is_ext)
        & content_ns.str.contains("납부허가", na=False)
        & content_ns.str.contains("세", na=False)
        & block
    ).map(lambda x: "Y" if x else "")
    df["신탁담보변경"] = (
        (~is_ext)
        & content_ns.str.contains("신탁|특약", na=False)
        & content_ns.str.contains("허가|의견서", na=False)
        & block
    ).map(lambda x: "Y" if x else "")

    df["열람대상"] = df[FLAG_COLS].eq("Y").any(axis=1).map(lambda x: "Y" if x else "")

    df["당부사항"] = ""
    mask_att = (
        df["열람대상"].eq("Y")
        & (
            content_ns.str.contains("신청서접수", na=False)
            | content_ns.str.contains("시부인표", na=False)
            | content_ns.str.contains("조사보고서", na=False)
        )
    )
    df.loc[mask_att, "당부사항"] = "별지 포함"

    mask_hg = df["심문기일"].eq("Y")
    df.loc[mask_hg, "당부사항"] = df.loc[mask_hg, "당부사항"].apply(
        lambda x: "심문조서 열람 요청" if not x else f"{x} / 심문조서 열람 요청"
    )
    mask_sv = df["현장검증기일"].eq("Y")
    df.loc[mask_sv, "당부사항"] = df.loc[mask_sv, "당부사항"].apply(
        lambda x: "현장검증조서 열람 요청" if not x else f"{x} / 현장검증조서 열람 요청"
    )

    df["_cns"] = content_ns
    df["_rns"] = result_ns
    df = df.drop_duplicates(subset=["사건번호", "일자", "_cns", "_rns"])
    df = df.drop(columns=["_cns", "_rns"])

    return df


def make_sheet_name(court, case_str):
    name = f"{court}_{case_str}"
    for ch in ['\\', '/', '*', '?', ':', '[', ']']:
        name = name.replace(ch, '')
    return name[:31]


def _ns(s):
    return re.sub(r"\s+", "", str(s)) if s else ""


def extract_key_dates(progress):
    rows = []
    for p in progress:
        일자 = p[0] if len(p) > 0 else ""
        내용 = p[1] if len(p) > 1 else ""
        결과 = p[2] if len(p) > 2 else ""
        if 일자:
            rows.append((str(일자).strip(), _ns(내용), _ns(결과)))

    def pick_first(cond):
        for 일자, c, r in rows:
            if cond(c, r):
                return 일자
        return "미확인"

    보전 = pick_first(lambda c, r: "보전처분" in c and "결정" in c
                      and "신청" not in c and "에게" not in c and "송달" not in c)

    개시 = pick_first(lambda c, r: "개시결정" in c and "관리인선임" in c
                      and "에게" not in c and "송달" not in c and "공고" not in c)
    if 개시 == "미확인":
        개시 = pick_first(lambda c, r: "개시결정" in c and "결정" in c
                          and "에게" not in c and "송달" not in c and "공고" not in c
                          and "신청" not in c and "의견서" not in c and "요청" not in c)

    시부인 = pick_first(lambda c, r: "시부인표" in c and "제출" in c
                        and "정정" not in c and "연장" not in c and "연기" not in c
                        and "에게" not in c)

    조사 = pick_first(lambda c, r: "조사보고서" in c and "제출" in c
                      and "연장" not in c and "연기" not in c and "기한" not in c
                      and "기간" not in c and "에게" not in c and "신청" not in c
                      and "명령" not in c)

    계획 = pick_first(lambda c, r: "회생계획안" in c and "제출" in c
                      and "기간" not in c and "연장" not in c and "에게" not in c
                      and "요지" not in c and "지정" not in c and "신청" not in c)

    인가폐지, 상태 = _extract_disposition(rows)

    return {
        "재산보전처분일": 보전,
        "개시결정일":     개시,
        "시부인표 제출일": 시부인,
        "조사보고서 제출일": 조사,
        "회생계획안 제출일": 계획,
        "인가폐지일":     인가폐지,
        "인가폐지상태":   상태,
    }


def _extract_disposition(rows):
    for 일자, c, r in rows:
        if c.startswith("종국:"):
            status = c.replace("종국:", "")
            if "취하" in status:
                status = "취하"
            return 일자, status

    def is_court_decision(c):
        return ("제출" not in c and "신청" not in c and "에게" not in c
                and "등기소" not in c and "등기관" not in c and "공고" not in c)

    for kw, label in [("인가결정", "인가"), ("폐지결정", "폐지"),
                      ("종결결정", "종결"), ("기각결정", "기각"),
                      ("각하결정", "각하"), ("취하허가결정", "취하"), ("취하", "취하")]:
        for 일자, c, r in rows:
            if kw in c and is_court_decision(c):
                return 일자, label

    return "미확인", ""


def _parse_date_tuple(s):
    m = re.findall(r'\d+', str(s))
    return tuple(int(x) for x in m[:3]) if len(m) >= 3 else (0, 0, 0)


def _clean_terminal_status(s):
    """종국결과를 총괄표 C열 표준값으로 정리한다."""
    raw = str(s or "").strip()
    t = _ns(raw)
    if not t:
        return ""
    t = re.sub(r"^종국[:：]?", "", t)
    if "기각" in t:
        return "기각"
    if "각하" in t:
        return "각하"
    if "취하" in t:
        return "취하"
    if "폐지" in t:
        return "폐지"
    if "종결" in t:
        return "종결"
    return raw


def _is_non_decision_noise(content_ns):
    """법원 결정 자체가 아닌 신청/제출/송달성 이벤트인지 판단한다."""
    noise_tokens = [
        "신청서", "의견서", "요청", "송달", "도달", "발송",
        "에게", "제출", "보정", "항고장", "허가신청"
    ]
    return any(t in content_ns for t in noise_tokens)


def _has_decision_keyword(content_ns, keyword):
    """신청/제출/송달 문구를 배제하고 결정 키워드 존재 여부를 판단한다."""
    return keyword in content_ns and not _is_non_decision_noise(content_ns)


def derive_progress_status(progress, basic=None):
    """총괄표 C열용 진행상황 도출.
    우선순위: 1.종국결과 2.인가결정 3.개시결정 4.개시신청"""
    basic = basic or {}

    terminal_from_basic = _clean_terminal_status(basic.get("종국결과", ""))
    if terminal_from_basic:
        return terminal_from_basic

    rows = []
    for p in progress or []:
        일자 = p[0] if len(p) > 0 else ""
        내용 = p[1] if len(p) > 1 else ""
        결과 = p[2] if len(p) > 2 else ""
        rgb  = p[4] if len(p) > 4 else "rgb(0,0,0)"

        c = _ns(내용)
        r = _ns(결과)
        if not c and not r:
            continue

        rows.append({
            "일자": str(일자 or "").strip(),
            "내용": str(내용 or "").strip(),
            "내용_ns": c,
            "결과_ns": r,
            "분류": categorize_color(rgb),
        })

    if not rows:
        return "미확인"

    for row in reversed(rows):
        c = row["내용_ns"]
        if c.startswith("종국:") or c.startswith("종국：") or c.startswith("종국"):
            terminal = _clean_terminal_status(c)
            if terminal:
                return terminal

    terminal_patterns = [
        ("기각결정", "기각"),
        ("각하결정", "각하"),
        ("취하허가결정", "취하"),
        ("폐지결정", "폐지"),
        ("종결결정", "종결"),
        ("기각", "기각"),
        ("각하", "각하"),
    ]
    for row in reversed(rows):
        c = row["내용_ns"]
        for kw, label in terminal_patterns:
            if _has_decision_keyword(c, kw):
                return label

    for row in reversed(rows):
        c = row["내용_ns"]
        if "회생계획" in c and _has_decision_keyword(c, "인가결정"):
            return "인가결정"

    for row in reversed(rows):
        c = row["내용_ns"]
        if _has_decision_keyword(c, "개시결정"):
            return "개시결정"

    return "개시신청"


def build_summary_sheet(wb, cases, results):
    ws = wb.create_sheet(title=SUMMARY_SHEET_NAME, index=0)

    headers = ["차주일련번호", "차주명(실명)", "진행상황", "관할법원", "회생사건번호",
               "신청일", "재산보전처분일", "개시결정일",
               "시부인표 제출일", "조사보고서 제출일", "회생계획안 제출일",
               "인가일 / 폐지(종결)일"]

    header_fill = PatternFill(fill_type="solid", fgColor="999999")
    header_font = Font(name="맑은 고딕", size=9, bold=True, color="FFFFFF")
    body_font   = Font(name="맑은 고딕", size=9)
    warn_font   = Font(name="맑은 고딕", size=9, color="C00000")
    center      = Alignment(horizontal="center", vertical="center")
    thin = Side(style="thin", color="BFBFBF")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for c, h in enumerate(headers, 1):
        cell = ws.cell(1, c, h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center
        cell.border = border

    meta = {}
    for cs in cases:
        meta[cs["사건번호"]] = (cs.get("일련번호", ""), cs.get("차주명", ""))

    row_idx = 2
    for r in results:
        case_str = r.get("사건번호", "")
        court    = r.get("관할법원", "")
        basic    = r.get("기본내용", {})
        progress = r.get("진행내용", [])
        serial, name = meta.get(case_str, ("", ""))

        신청일 = ""
        if progress:
            신청일 = str(progress[0][0]).strip() if progress[0] and progress[0][0] else ""
        if not 신청일:
            신청일 = basic.get("접수일", "")
        if not 신청일:
            신청일 = "미확인"

        dates = extract_key_dates(progress)
        진행상황 = derive_progress_status(progress, basic)
        인가폐지 = dates["인가폐지일"]
        if 인가폐지 not in ("", "미확인") and dates["인가폐지상태"]:
            인가폐지 = f'{인가폐지} ({dates["인가폐지상태"]})'

        values = [
            serial, name, 진행상황, court, case_str,
            신청일,
            dates["재산보전처분일"],
            dates["개시결정일"],
            dates["시부인표 제출일"],
            dates["조사보고서 제출일"],
            dates["회생계획안 제출일"],
            인가폐지,
        ]

        for c, v in enumerate(values, 1):
            cell = ws.cell(row_idx, c, v)
            cell.alignment = center
            cell.border = border
            if v == "미확인":
                cell.font = warn_font
            else:
                cell.font = body_font

        row_idx += 1

    widths = {"A": 11, "B": 22, "C": 20, "D": 14, "E": 15, "F": 13,
              "G": 14, "H": 13, "I": 14, "J": 15, "K": 15, "L": 20}
    for col, w in widths.items():
        ws.column_dimensions[col].width = w

    ws.freeze_panes = "A2"
    return ws


def save_results(results, path, cases=None):
    if not results:
        return

    rows = []
    for r in results:
        basic    = r.get("기본내용", {})
        progress = r.get("진행내용", [])
        case_str = r.get("사건번호", "")
        court    = r.get("관할법원", "")

        if not progress:
            rows.append({
                "사건번호": case_str, "관할법원": court,
                "일자": "", "내용": "", "결과": "", "공시문": "",
                "내용_rgb": "rgb(0,0,0)",
                "조회결과": r.get("결과", ""), "오류내용": r.get("오류내용", ""),
                "조회일시": r.get("조회일시", ""),
                **{k: basic.get(k, "") for k in ["사건명", "신청인", "채권자", "재판부", "접수일", "종국결과", "채무자"]},
            })
        else:
            for pd_row in progress:
                rows.append({
                    "사건번호": case_str, "관할법원": court,
                    "일자":   pd_row[0] if len(pd_row) > 0 else "",
                    "내용":   pd_row[1] if len(pd_row) > 1 else "",
                    "결과":   pd_row[2] if len(pd_row) > 2 else "",
                    "공시문": pd_row[3] if len(pd_row) > 3 else "",
                    "내용_rgb": pd_row[4] if len(pd_row) > 4 else "rgb(0,0,0)",
                    "조회결과": r.get("결과", ""), "오류내용": r.get("오류내용", ""),
                    "조회일시": r.get("조회일시", ""),
                    **{k: basic.get(k, "") for k in ["사건명", "신청인", "채권자", "재판부", "접수일", "종국결과", "채무자"]},
                })

    df = pd.DataFrame(rows)
    df = compute_flags(df)

    df["분류"] = df["내용_rgb"].apply(categorize_color)

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    label_fill        = PatternFill(fill_type="solid", fgColor="D9E1F2")
    label_font        = Font(bold=True, name="Arial", size=9)
    header_fill       = PatternFill(fill_type="solid", fgColor="2F5496")
    header_font_white = Font(bold=True, color="FFFFFF", name="Arial", size=9)
    center            = Alignment(horizontal="center", vertical="center")
    wrap              = Alignment(wrap_text=True, vertical="top")

    HEADER_ROW_IDX = 8

    for r in results:
        case_str = r.get("사건번호", "")
        court    = r.get("관할법원", "")
        basic    = r.get("기본내용", {})

        sheet_name = make_sheet_name(court, case_str)
        ws = wb.create_sheet(title=sheet_name)

        rows_basic = [
            [("사건번호", basic.get("사건번호", case_str)), ("사건명",   basic.get("사건명", ""))],
            [("신청인",   basic.get("신청인", basic.get("채무자", ""))), ("채권자",   basic.get("채권자", ""))],
            [("재판부",   basic.get("재판부", "")),                       ("관할법원", court)],
            [("접수일",   basic.get("접수일", "")),                       ("종국결과", basic.get("종국결과", ""))],
        ]
        for ri, pairs in enumerate(rows_basic, 1):
            for ci_pair, (label, value) in enumerate(pairs):
                ci = ci_pair * 2 + 1
                cl = ws.cell(ri, ci, label)
                cl.font = label_font
                cl.fill = label_fill
                cl.alignment = center
                cv = ws.cell(ri, ci + 1, value)
                cv.font = Font(name="Arial", size=9)

        for ci_pair, (label, value) in enumerate([
            ("조회일시", r.get("조회일시", "")),
            ("조회결과", r.get("결과", ""))
        ]):
            ci = ci_pair * 2 + 1
            cl = ws.cell(5, ci, label)
            cl.font = label_font
            cl.fill = label_fill
            cl.alignment = center
            cv = ws.cell(5, ci + 1, value)
            cv.font = Font(name="Arial", size=9)

        all_headers = ["일자", "내용", "결과", "공시문"] + FLAG_HEADER_COLS
        for ci, h in enumerate(all_headers, 1):
            cell = ws.cell(HEADER_ROW_IDX, ci, h)
            cell.font = header_font_white
            cell.fill = header_fill
            cell.alignment = center

        case_df = df[df["사건번호"] == case_str].reset_index(drop=True)

        for ri, row in case_df.iterrows():
            excel_row = HEADER_ROW_IDX + 1 + ri
            rgb_str   = row.get("내용_rgb", "rgb(0,0,0)")
            hex_color = rgb_to_hex(rgb_str)

            for ci, col in enumerate(["일자", "내용", "결과", "공시문"], 1):
                c = ws.cell(excel_row, ci, row.get(col, ""))
                c.alignment = wrap
                c.font = Font(name="Arial", size=9, color=hex_color)

            for ci, col in enumerate(FLAG_HEADER_COLS, 5):
                c = ws.cell(excel_row, ci, row.get(col, ""))
                c.alignment = center
                c.font = Font(name="Arial", size=9)

        ws.column_dimensions["A"].width = 14
        ws.column_dimensions["B"].width = 55
        ws.column_dimensions["C"].width = 22
        ws.column_dimensions["D"].width = 12
        ws.column_dimensions["E"].width = 10
        for ci in range(6, len(all_headers) + 1):
            col_letter = openpyxl.utils.get_column_letter(ci)
            ws.column_dimensions[col_letter].width = 10
        for ci in range(8, len(all_headers) + 1):
            col_letter = openpyxl.utils.get_column_letter(ci)
            ws.column_dimensions[col_letter].hidden = True

        ws.auto_filter.ref = f"A{HEADER_ROW_IDX}:{openpyxl.utils.get_column_letter(len(all_headers))}{HEADER_ROW_IDX}"

    if cases is not None:
        build_summary_sheet(wb, cases, results)

    wb.save(path)
    print(f"저장 완료: {path}")


def run_crawl(cases, api_key, btpr_nm, output_path,
              on_log=None, on_progress=None, on_case_done=None,
              should_stop=None, delay=None):
    def log(msg):
        if on_log:
            on_log(msg)
        else:
            print(msg)

    if delay is None:
        delay = DELAY_BETWEEN_REQUESTS

    total = len(cases)
    log("=" * 55)
    log("  대법원 회생사건 자동 조회 v11.2")
    log("=" * 55)

    if not api_key:
        log("❌ 인증 API 키가 없습니다.")
        return [], {"ok": 0, "none": 0, "err": 0, "total": total, "path": None}

    log(f"✅ API 키 확인 완료 ({api_key[:6]}...)")
    log(f"📋 조회 대상: 총 {total}건 / 당사자명: '{btpr_nm}'")
    log("🌐 브라우저 실행 중...")

    driver = setup_driver()
    results = []

    try:
        for i, case in enumerate(cases, 1):
            if should_stop and should_stop():
                log("\n⏹ 사용자 중단")
                break

            if on_progress:
                on_progress(i - 1, total)

            log(f"\n[{i}/{total}] {case['법원']} / {case['사건번호']}")
            t0 = time.time()
            r = search_one(driver, api_key, case["법원"], case["사건번호"], btpr_nm=btpr_nm)
            elapsed = time.time() - t0
            results.append(r)

            prog_count = len(r.get("진행내용", []))
            결과 = r.get("결과", "")
            오류 = r.get("오류내용", "")
            log(f"  → {결과} {오류} | 진행내용 {prog_count}건 | {elapsed:.1f}초")

            if on_case_done:
                on_case_done(i, total, r, elapsed)
            if on_progress:
                on_progress(i, total)

            if i % 5 == 0:
                save_results(results, output_path, cases=cases)
                log(f"  💾 중간 저장 ({i}건)")

            time.sleep(delay)

    except KeyboardInterrupt:
        log("\n⏹ 중단됨 (Ctrl+C)")
    finally:
        try:
            driver.quit()
        except Exception:
            pass

    save_results(results, output_path, cases=cases)

    ok   = sum(1 for r in results if r.get("결과") == "조회성공")
    none = sum(1 for r in results if r.get("결과") == "사건 없음")
    err  = len(results) - ok - none
    log(f"\n완료 — 성공: {ok}건 / 사건없음: {none}건 / 오류: {err}건")
    log(f"결과파일: {output_path}")

    return results, {"ok": ok, "none": none, "err": err, "total": total, "path": output_path}


def main():
    if not VERIFY_API_KEY:
        print("❌ .env_verify 파일에 VERIFY_API_KEY가 없습니다.")
        return
    print(f"\n엑셀 로드: {INPUT_EXCEL_PATH}")
    try:
        cases = load_cases(INPUT_EXCEL_PATH, INPUT_SHEET_NAME,
                           COL_SERIAL, COL_NAME, COL_COURT, COL_CASE_NUMBER, HEADER_ROW)
    except FileNotFoundError:
        print(f"❌ 파일 없음: {INPUT_EXCEL_PATH}")
        return
    run_crawl(cases, VERIFY_API_KEY, BTPR_NM, OUTPUT_EXCEL_PATH)


if __name__ == "__main__":
    main()
