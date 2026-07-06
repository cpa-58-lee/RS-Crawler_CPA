"""
대법원 회생사건 자동 조회 크롤러 v10
- 자동입력방지문자(보안문자) 자동 해제
- 진행내용 전체 파싱 + 글자색 반영
- 플래그 자동 분류 (열람대상, 당부사항 등 13개 항목)
- 조회일시/결과 5행 표시
- 맨 앞 "회생차주 총괄표" 시트 자동 생성 (주요 일정 자동 추출)
- 파산(하합/하단)·개인회생(개회) 사건번호 지원
- alert/팝업 대기 최적화로 사건당 속도 단축
- 플래그 로직에 "연기"/"수정" 제외 조건
- 진행내용 수집 실패 시 자동 재시도 (성공 오기록 방지)
설치: pip install selenium openpyxl 2captcha-python webdriver-manager python-dotenv pandas
.env_verify: VERIFY_API_KEY=실제키입력
"""

import time
import re
import os
from datetime import datetime
from dotenv import load_dotenv
from selenium import webdriver
from selenium.common.exceptions import NoAlertPresentException, TimeoutException
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import Select, WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.chrome.service import Service
from webdriver_manager.chrome import ChromeDriverManager
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import pandas as pd
from twocaptcha import TwoCaptcha as _VerifySolver

load_dotenv(dotenv_path=".env_verify")
# 하위호환: 구 파일명/변수도 함께 확인
VERIFY_API_KEY = os.getenv("VERIFY_API_KEY", "") or os.getenv("TWOCAPTCHA_API_KEY", "")
if not VERIFY_API_KEY:
    try:
        from dotenv import dotenv_values
        _vals = {**dotenv_values(".env_verify"), **dotenv_values(".env_capcha")}
        VERIFY_API_KEY = _vals.get("VERIFY_API_KEY", "") or _vals.get("TWOCAPTCHA_API_KEY", "")
    except Exception:
        pass

# ============================================================
# ✅ 매 딜마다 여기만 수정하세요
# ============================================================
BTPR_NM = "아이"   # 당사자명: 농협, 신한, 국민, 하나, 우리 등

INPUT_EXCEL_PATH  = r"C:\Users\LEE CPA\OneDrive\Desktop\Record\python Project\5. 나의사건검색 크롤링\INPUT DATA.xlsx"
INPUT_SHEET_NAME  = "Sheet1"
COL_SERIAL        = "A"   # ★ v10: 차주일련번호 (EY-XXX)
COL_NAME          = "B"   # ★ v10: 차주명(실명)
COL_COURT         = "C"
COL_CASE_NUMBER   = "D"
HEADER_ROW        = 1
OUTPUT_EXCEL_PATH = f"회생사건_조회결과_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
DELAY_BETWEEN_REQUESTS = 1.0   # ★ v10: 2.0 → 1.0
MAX_CAPTCHA_RETRIES    = 5

SUMMARY_SHEET_NAME = "회생차주 총괄표"   # ★ v10

TARGET_URL     = "https://ssgo.scourt.go.kr/ssgo/index.on?cortId=www"
ID_COURT       = "mf_ssgoTopMainTab_contents_content1_body_sbx_cortCd"
ID_FULL_CS_NO  = "mf_ssgoTopMainTab_contents_content1_body_ibx_fullCsNo"
ID_BTPR_NM     = "mf_ssgoTopMainTab_contents_content1_body_ibx_btprNm"
ID_SANO_MODE   = "mf_ssgoTopMainTab_contents_content1_body_cbx_chkSanoInputMode_input_0"
ID_CAPTCHA_IMG = "mf_ssgoTopMainTab_contents_content1_body_img_captcha"
ID_CAPTCHA_INP = "mf_ssgoTopMainTab_contents_content1_body_ibx_answer"
ID_SEARCH_BTN  = "mf_ssgoTopMainTab_contents_content1_body_btn_srchCs"
ID_TAB2        = "mf_ssgoTopMainTab_contents_content1_body_wfSsgoDetail_ssgoCsDetailTab_tab_ssgoTab2_tabHTML"


FLAG_COLS = [
    "신청서접수", "답변서제출", "심문사항답변", "심문기일", "현장검증기일",
    "급여", "시부인표", "채권신고및변제요청서", "월간보고서", "기타보고서",
    "조사보고서", "회생계획안", "임금채권공익채권",
]

FLAG_HEADER_COLS = [
    "당부사항", "열람대상",
    "신청서접수", "답변서제출", "심문사항답변", "심문기일", "현장검증기일",
    "급여", "시부인표", "채권신고및변제요청서", "월간보고서", "기타보고서",
    "조사보고서", "회생계획안", "임금채권공익채권",
]


# ============================================================
# 드라이버
# ============================================================
def setup_driver():
    options = webdriver.ChromeOptions()
    options.add_argument("--no-sandbox")
    options.add_argument("--disable-dev-shm-usage")
    options.add_argument("--window-size=1280,900")
    options.add_argument("user-agent=Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36")
    driver = webdriver.Chrome(service=Service(ChromeDriverManager().install()), options=options)
    driver.set_page_load_timeout(30)
    return driver


# ============================================================
# 붙여넣기 파싱 (GUI용)  ★ v10
# ============================================================
_HEADER_KEYWORDS = ["일련번호", "차주일련번호", "차주명", "관할법원", "회생사건번호", "사건번호"]
_CASE_HINT = re.compile(r'\d{4}(간회합|간회단|회합|회단|간회|개회|하합|하단|간이|하)\d+')


def _looks_like_header(cells):
    """헤더 행 판정: 사건번호 패턴이 없고 헤더 키워드가 셀 전체로 2개 이상이면 헤더."""
    joined = "".join(cells).replace(" ", "")
    if _CASE_HINT.search(joined):
        return False
    keyset = [k.replace(" ", "") for k in _HEADER_KEYWORDS]
    hits = sum(1 for c in cells if c.replace(" ", "") in keyset)
    return hits >= 2


def parse_pasted_table(text):
    """엑셀에서 복사한 표(탭 구분) → cases 리스트.
    컬럼 순서: 일련번호 / 차주명 / 법원 / 사건번호. 헤더/빈줄 자동 처리."""
    cases = []
    lines = text.replace("\r\n", "\n").replace("\r", "\n").split("\n")
    for ln in lines:
        if not ln.strip():
            continue
        if "\t" in ln:
            cells = [c.strip() for c in ln.split("\t")]
        elif "," in ln:
            cells = [c.strip() for c in ln.split(",")]
        else:
            cells = [c.strip() for c in re.split(r"\s{2,}", ln)]
        if len(cells) < 4:
            continue
        if _looks_like_header(cells):
            continue
        serial, name, court, case = cells[0], cells[1], cells[2], cells[3]
        if not case or case.lower() == "none":
            continue
        cases.append({"일련번호": serial, "차주명": name, "법원": court, "사건번호": case})
    return cases


# ============================================================
# 엑셀 로드  ★ v10: 일련번호/차주명 컬럼 추가
# ============================================================
def load_cases(excel_path, sheet_name, col_serial, col_name, col_court, col_case, header_row):
    wb = openpyxl.load_workbook(excel_path, data_only=True)
    ws = wb[sheet_name]
    ci_serial = openpyxl.utils.column_index_from_string(col_serial) - 1
    ci_name   = openpyxl.utils.column_index_from_string(col_name)   - 1
    ci_court  = openpyxl.utils.column_index_from_string(col_court)  - 1
    ci_case   = openpyxl.utils.column_index_from_string(col_case)   - 1
    cases = []
    for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
        serial = str(row[ci_serial]).strip() if row[ci_serial] else ""
        name   = str(row[ci_name]).strip()   if row[ci_name]   else ""
        court  = str(row[ci_court]).strip()  if row[ci_court]  else ""
        case   = str(row[ci_case]).strip()   if row[ci_case]   else ""
        if case and case != "None":
            cases.append({"일련번호": serial, "차주명": name, "법원": court, "사건번호": case})
    print(f"총 {len(cases)}건 로드 완료")
    return cases


# ============================================================
# 유틸
# ============================================================
# ★ v10: 회생(회합/회단/간회합/간회단) + 파산(하합/하단/간이) + 개인회생(개회) 지원
CASE_TYPE_RE = re.compile(
    r'(\d{4})'
    r'(간회합|간회단|회합|회단'      # 법인/일반 회생
    r'|간회|개회'                    # 개인회생 (개회) 및 간이회생 축약형
    r'|하합|하단|간이|하'            # 파산 (하합/하단 등)
    r')'
    r'(\d+)$'
)


def parse_case_number(s):
    m = CASE_TYPE_RE.search(s.strip())
    if m:
        return {"year": m.group(1), "type_code": m.group(2), "serial": m.group(3)}
    return None


def rgb_to_hex(rgb_str):
    try:
        nums = re.findall(r'\d+', rgb_str)
        return '{:02X}{:02X}{:02X}'.format(int(nums[0]), int(nums[1]), int(nums[2]))
    except:
        return '000000'


def wait_el(driver, el_id, timeout=10):
    return WebDriverWait(driver, timeout).until(EC.presence_of_element_located((By.ID, el_id)))


def quick_accept_alert(driver, timeout=0.4):
    """★ v10: alert 있으면 즉시 닫고 텍스트 반환, 없으면 짧게 폴링 후 바로 리턴.
    기존 1.5~2초 블로킹을 0.4초 폴링으로 대체."""
    end = time.time() + timeout
    while time.time() < end:
        try:
            alert = driver.switch_to.alert
            txt = alert.text
            alert.accept()
            return txt
        except NoAlertPresentException:
            time.sleep(0.05)
    return None


def dismiss_popup(driver, timeout=0.4):
    """★ v10: '확인' 버튼 팝업 즉시 처리, 없으면 짧게 폴링 후 리턴."""
    end = time.time() + timeout
    while time.time() < end:
        try:
            btns = driver.find_elements(By.XPATH, "//button[contains(text(),'확인')]")
            for btn in btns:
                if btn.is_displayed() and btn.is_enabled():
                    btn.click()
                    return True
        except:
            pass
        time.sleep(0.05)
    return False


def select_court(driver, court, timeout=15):
    try:
        sel = driver.find_element(By.ID, ID_COURT)
        if len(Select(sel).options) <= 200:
            WebDriverWait(driver, timeout).until(
                lambda d: len(Select(d.find_element(By.ID, ID_COURT)).options) > 200
            )
    except:
        WebDriverWait(driver, timeout).until(
            lambda d: len(Select(d.find_element(By.ID, ID_COURT)).options) > 200
        )
    driver.execute_script("""
        var sel = document.getElementById(arguments[0]);
        for (var i = 0; i < sel.options.length; i++) {
            if (sel.options[i].value === arguments[1] || sel.options[i].text === arguments[1]) {
                sel.selectedIndex = i;
                sel.dispatchEvent(new Event('change', {bubbles: true}));
                return;
            }
        }
    """, ID_COURT, court)


def js_set_input(driver, el_id, value):
    """JS로 input 값 세팅 (native setter 방식 - validation 우회)"""
    driver.execute_script("""
        var el = document.getElementById(arguments[0]);
        el.focus();
        var nativeInputValueSetter = Object.getOwnPropertyDescriptor(
            window.HTMLInputElement.prototype, 'value').set;
        nativeInputValueSetter.call(el, arguments[1]);
        el.dispatchEvent(new Event('input', {bubbles: true}));
        el.dispatchEvent(new Event('change', {bubbles: true}));
        el.dispatchEvent(new KeyboardEvent('keyup', {bubbles: true}));
    """, el_id, value)


# ============================================================
# 보안문자(자동입력방지) 처리
# ============================================================
def get_captcha_base64(driver):
    try:
        WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.ID, ID_CAPTCHA_IMG)))
        return driver.execute_script("""
            var img = document.getElementById(arguments[0]);
            if (!img) return null;
            var canvas = document.createElement('canvas');
            canvas.width  = img.naturalWidth  || img.width  || 130;
            canvas.height = img.naturalHeight || img.height || 50;
            var ctx = canvas.getContext('2d');
            ctx.drawImage(img, 0, 0);
            return canvas.toDataURL('image/png').split(',')[1];
        """, ID_CAPTCHA_IMG)
    except Exception as e:
        print(f"  보안문자 이미지 추출 실패: {e}")
        return None


def resolve_verification(api_key, image_b64):
    """자동입력방지문자(보안문자) 해제."""
    solver = _VerifySolver(api_key)
    try:
        result = solver.normal(image_b64)
        return result["code"]
    except Exception as e:
        print(f"  인증 처리 실패: {e}")
        return None


# ============================================================
# 결과 파싱
# ============================================================
def parse_result(driver, court, case_str):
    try:
        # alert 팝업 먼저 처리 (보안문자 오답 alert)
        alert_text = quick_accept_alert(driver, timeout=0.6)
        if alert_text and ("자동입력" in alert_text or "방지문자" in alert_text):
            return {"_retry": True}

        src = driver.page_source

        if any(x in src for x in ["자동입력방지문자가 올바르지", "보안문자를 다시", "captchaError"]):
            return {"_retry": True}

        if any(x in src for x in ["검색된 사건이 없습니다", "해당 사건이 없습니다", "조회된 내역이 없"]):
            return {"사건번호": case_str, "관할법원": court, "결과": "사건 없음",
                    "조회일시": datetime.now().strftime("%Y-%m-%d %H:%M"), "기본내용": {}, "진행내용": []}

        # 진행내용 탭 클릭 (여러 번 재시도 + 데이터 행 채워질 때까지 대기)
        tab_ok = False
        for tab_try in range(3):
            try:
                WebDriverWait(driver, 8).until(
                    EC.presence_of_element_located((By.ID, ID_TAB2))
                )
                driver.execute_script("document.getElementById(arguments[0]).click();", ID_TAB2)
                if _wait_progress_table(driver, timeout=8.0):
                    tab_ok = True
                    break
                time.sleep(0.5)
            except Exception as e:
                if tab_try == 2:
                    print(f"  진행내용 탭 클릭 실패(3회): {e}")
                time.sleep(0.6)

        # 기본내용 파싱
        basic = {"관할법원": court}
        BASIC_FIELDS = ["사건번호", "사건명", "신청인", "채권자", "재판부", "접수일", "종국결과", "채무자"]
        try:
            for row in driver.find_elements(By.CSS_SELECTOR, "table tr"):
                ths = row.find_elements(By.TAG_NAME, "th")
                tds = row.find_elements(By.TAG_NAME, "td")
                for th, td in zip(ths, tds):
                    k = th.text.strip()
                    v = td.text.strip()
                    if k in BASIC_FIELDS and v and k not in basic:
                        basic[k] = v
        except Exception as e:
            print(f"  기본내용 파싱 실패: {e}")

        # 진행내용 테이블 파싱 - JS 한 번에 전체 추출
        progress = []
        try:
            for table in driver.find_elements(By.CSS_SELECTOR, "table"):
                headers = [th.text.strip() for th in table.find_elements(By.TAG_NAME, "th")]
                if "일자" in headers and "내용" in headers and "결과" in headers and "공시문" in headers:
                    rows_data = driver.execute_script("""
                        var rows = arguments[0].querySelectorAll('tr');
                        var result = [];
                        for (var i = 1; i < rows.length; i++) {
                            var cells = rows[i].querySelectorAll('td');
                            if (cells.length < 2) continue;
                            var color = window.getComputedStyle(cells[1]).color;
                            result.push([
                                cells[0] ? cells[0].innerText.trim() : '',
                                cells[1] ? cells[1].innerText.trim() : '',
                                cells[2] ? cells[2].innerText.trim() : '',
                                cells[3] ? cells[3].innerText.trim() : '',
                                color
                            ]);
                        }
                        return result;
                    """, table)
                    for row_data in rows_data:
                        if row_data[0] or row_data[1]:
                            progress.append(row_data)
                    break
        except Exception as e:
            print(f"  진행내용 파싱 실패: {e}")

        # ★ 진행내용 수집 검증
        # 위에서 "사건 없음"은 이미 걸러졌으므로, 여기서 진행내용 0건이면 수집 실패로 판단.
        # (탭이 열렸더라도 렌더링 지연 등으로 데이터를 못 가져온 경우 → 재시도)
        if not progress:
            return {"_retry_fetch": True}

        return {"사건번호": case_str, "관할법원": court, "결과": "조회성공",
                "조회일시": datetime.now().strftime("%Y-%m-%d %H:%M"),
                "기본내용": basic, "진행내용": progress}

    except Exception as e:
        return {"사건번호": case_str, "관할법원": court, "결과": "오류", "오류내용": str(e),
                "조회일시": datetime.now().strftime("%Y-%m-%d %H:%M"), "기본내용": {}, "진행내용": []}


def _wait_progress_table(driver, timeout=8.0):
    """진행내용 테이블의 '데이터 행'이 실제로 채워지고 행 수가 안정될 때까지 대기.
    (헤더만 확인하던 기존 방식은 렌더링 중간에 통과되어 0건 누락 발생 → 데이터 행 기준으로 변경)"""
    end = time.time() + timeout
    last_count = -1
    stable = 0
    while time.time() < end:
        try:
            count = driver.execute_script("""
                var tables = document.querySelectorAll('table');
                for (var t = 0; t < tables.length; t++) {
                    var ths = tables[t].querySelectorAll('th');
                    var txt = '';
                    for (var i = 0; i < ths.length; i++) txt += ths[i].innerText;
                    if (txt.indexOf('일자') >= 0 && txt.indexOf('내용') >= 0 &&
                        txt.indexOf('결과') >= 0 && txt.indexOf('공시문') >= 0) {
                        // 데이터 행(td 2개 이상 가진 tr) 개수
                        var rows = tables[t].querySelectorAll('tr');
                        var n = 0;
                        for (var r = 1; r < rows.length; r++) {
                            if (rows[r].querySelectorAll('td').length >= 2) n++;
                        }
                        return n;
                    }
                }
                return -1;   // 테이블 자체가 아직 없음
            """)
            if count is not None and count >= 1:
                # 행 수가 연속 2회 동일하면 렌더링 완료로 판단
                if count == last_count:
                    stable += 1
                    if stable >= 2:
                        return True
                else:
                    stable = 0
                last_count = count
        except:
            pass
        time.sleep(0.15)
    # 타임아웃: 데이터 행이 1개라도 잡혔으면 True, 아니면 False
    return last_count >= 1


# ============================================================
# 검색  ★ v10: alert 대기 최적화
# ============================================================
def search_one(driver, api_key, court, case_str, btpr_nm=None):
    if btpr_nm is None:
        btpr_nm = BTPR_NM
    parsed = parse_case_number(case_str)
    if not parsed:
        return {"사건번호": case_str, "관할법원": court, "결과": "오류",
                "오류내용": "사건번호 형식 인식 불가",
                "조회일시": datetime.now().strftime("%Y-%m-%d %H:%M"), "기본내용": {}, "진행내용": []}

    for attempt in range(1, MAX_CAPTCHA_RETRIES + 1):
        try:
            driver.get(TARGET_URL)
            WebDriverWait(driver, 15).until(
                EC.presence_of_element_located((By.ID, ID_SANO_MODE))
            )
            # ★ v10: 잔여 alert 즉시 제거 (블로킹 없음)
            quick_accept_alert(driver, timeout=0.2)

            sano_chk = wait_el(driver, ID_SANO_MODE)
            if not sano_chk.is_selected():
                sano_chk.click()

            select_court(driver, court)

            # 사건번호 입력
            full_input = WebDriverWait(driver, 10).until(
                EC.element_to_be_clickable((By.ID, ID_FULL_CS_NO))
            )
            full_input.click()
            full_input.clear()
            full_input.send_keys(case_str)
            # ★ v10: alert/팝업 짧게 처리 (기존 1.5초×2 → 0.3초×2)
            quick_accept_alert(driver, timeout=0.3)
            dismiss_popup(driver, timeout=0.2)

            # 당사자명 입력 (JS 방식)
            js_set_input(driver, ID_BTPR_NM, btpr_nm)
            quick_accept_alert(driver, timeout=0.3)
            dismiss_popup(driver, timeout=0.2)

            print(f"  [{attempt}/{MAX_CAPTCHA_RETRIES}] 보안문자 확인 중...")
            b64 = get_captcha_base64(driver)
            if not b64:
                continue

            answer = resolve_verification(api_key, b64)
            if not answer:
                continue
            print(f"  [{attempt}/{MAX_CAPTCHA_RETRIES}] 보안문자: '{answer}'")

            # 보안문자 입력 (JS 방식)
            js_set_input(driver, ID_CAPTCHA_INP, answer)
            dismiss_popup(driver, timeout=0.2)

            driver.execute_script("document.getElementById(arguments[0]).click();", ID_SEARCH_BTN)
            # ★ v10: 고정 1.5초 대신 결과 or alert 등장까지 폴링
            _wait_after_search(driver, timeout=4.0)

            result = parse_result(driver, court, case_str)
            if result.get("_retry"):
                print(f"  [{attempt}/{MAX_CAPTCHA_RETRIES}] 자동입력 확인 재시도...")
                continue
            # ★ 진행내용 수집 실패 → 페이지 새로 열어 재시도
            if result.get("_retry_fetch"):
                print(f"  [{attempt}/{MAX_CAPTCHA_RETRIES}] 진행내용 수집 실패, 재시도...")
                time.sleep(1)
                continue

            return result

        except Exception as e:
            quick_accept_alert(driver, timeout=0.3)
            print(f"  [{attempt}/{MAX_CAPTCHA_RETRIES}] 예외: {e}")
            time.sleep(1)

    return {"사건번호": case_str, "관할법원": court, "결과": "오류",
            "오류내용": "진행내용 수집 실패 (재시도 초과)",
            "조회일시": datetime.now().strftime("%Y-%m-%d %H:%M"), "기본내용": {}, "진행내용": []}


def _wait_after_search(driver, timeout=4.0):
    """★ v10: 검색 클릭 후 (진행내용 탭 | 사건없음 | 보안문자 alert) 중 하나 등장까지만 대기."""
    end = time.time() + timeout
    while time.time() < end:
        # alert 먼저
        try:
            _ = driver.switch_to.alert
            return  # alert 있으면 parse_result가 처리
        except NoAlertPresentException:
            pass
        try:
            done = driver.execute_script("""
                if (document.getElementById(arguments[0])) return true;
                var b = document.body ? document.body.innerText : '';
                if (b.indexOf('검색된 사건이 없') >= 0 ||
                    b.indexOf('해당 사건이 없') >= 0 ||
                    b.indexOf('올바르지') >= 0) return true;
                return false;
            """, ID_TAB2)
            if done:
                return
        except:
            pass
        time.sleep(0.1)


# ============================================================
# 플래그 계산  ★ v10: "연기"/"수정" 제외 조건 추가
# ============================================================
def compute_flags(df: pd.DataFrame) -> pd.DataFrame:
    content_ns = df["내용"].fillna("").astype(str).str.replace(r"\s+", "", regex=True)
    result_ns  = df["결과"].fillna("").astype(str).str.replace(r"\s+", "", regex=True)

    is_ext = content_ns.str.contains("연장", na=False)
    # ★ v10: "연기" 또는 "수정" 포함 행은 플래그 제외
    is_defer_or_edit = content_ns.str.contains("연기|수정", na=False)

    block  = (
        ~content_ns.str.contains("송달|도달|변경", na=False)
        & ~result_ns.str.contains("송달|도달|변경", na=False)
        & ~is_defer_or_edit          # ★ v10
    )

    def flag_all(*tokens):
        cond = ~is_ext
        for t in tokens:
            cond &= content_ns.str.contains(t, na=False)
        return (cond & block).map(lambda x: "Y" if x else "")

    df["신청서접수"]           = flag_all("신청서접수")
    df["답변서제출"]           = flag_all("답변서", "제출")
    df["심문사항답변"]         = flag_all("심문사항")
    df["심문기일"]             = (
        (~is_ext) & content_ns.str.contains("심문기일", na=False)
        & content_ns.str.contains("심문실|법|:", na=False) & block
    ).map(lambda x: "Y" if x else "")
    df["현장검증기일"]         = (
        (~is_ext) & content_ns.str.contains("현장검증기일", na=False)
        & content_ns.str.contains(":", na=False) & block
    ).map(lambda x: "Y" if x else "")
    df["급여"]                 = flag_all("급여", "지급")
    df["시부인표"]             = flag_all("시부인표")
    df["채권신고및변제요청서"] = flag_all("채권신고및변제요청서")
    df["월간보고서"]           = flag_all("월간보고서")
    df["기타보고서"]           = (
        (~is_ext) & content_ns.str.contains("기타보고서", na=False)
        & content_ns.str.contains("제출", na=False) & block
    ).map(lambda x: "Y" if x else "")
    df["조사보고서"]           = flag_all("조사보고서")
    df["회생계획안"]           = (
        flag_all("회생계획안").eq("Y")
        & ~content_ns.str.contains("지정|연기|발송", na=False)
    ).map(lambda x: "Y" if x else "")
    df["임금채권공익채권"]     = flag_all("임금채권")

    # 열람대상
    df["열람대상"] = df[FLAG_COLS].eq("Y").any(axis=1).map(lambda x: "Y" if x else "")

    # 당부사항
    df["당부사항"] = ""
    mask_att = (
        df["열람대상"].eq("Y")
        & (
            content_ns.str.contains("신청서접수", na=False)
            | content_ns.str.contains("시부인표", na=False)
            | content_ns.str.contains("조사보고서", na=False)
        )
    )
    df.loc[mask_att, "당부사항"] = "별지 포함"

    mask_hg = df["심문기일"].eq("Y")
    df.loc[mask_hg, "당부사항"] = df.loc[mask_hg, "당부사항"].apply(
        lambda x: "심문조서 열람 요청" if not x else f"{x} / 심문조서 열람 요청"
    )
    mask_sv = df["현장검증기일"].eq("Y")
    df.loc[mask_sv, "당부사항"] = df.loc[mask_sv, "당부사항"].apply(
        lambda x: "현장검증조서 열람 요청" if not x else f"{x} / 현장검증조서 열람 요청"
    )

    # 중복 제거
    df["_cns"] = content_ns
    df["_rns"] = result_ns
    df = df.drop_duplicates(subset=["사건번호", "일자", "_cns", "_rns"])
    df = df.drop(columns=["_cns", "_rns"])

    return df


# ============================================================
# 저장
# ============================================================
def make_sheet_name(court, case_str):
    name = f"{court}_{case_str}"
    for ch in ['\\', '/', '*', '?', ':', '[', ']']:
        name = name.replace(ch, '')
    return name[:31]


# ★ v10: 진행내용에서 주요 날짜 확정 추출 (XLOOKUP 대체)
def _ns(s):
    """공백 제거 정규화"""
    return re.sub(r"\s+", "", str(s)) if s else ""


def extract_key_dates(progress):
    """진행내용(list of [일자,내용,결과,공시,rgb])에서 각 항목 확정일 1개 추출.
    매칭 없으면 '미확인' 반환.
    규칙은 실제 대법원 진행내용 문구 패턴 기반 (송달/신청/연장/등기소 노이즈 제거).
    """
    rows = []
    for p in progress:
        일자 = p[0] if len(p) > 0 else ""
        내용 = p[1] if len(p) > 1 else ""
        결과 = p[2] if len(p) > 2 else ""
        if 일자:
            rows.append((str(일자).strip(), _ns(내용), _ns(결과)))

    def pick_first(cond):
        for 일자, c, r in rows:
            if cond(c, r):
                return 일자
        return "미확인"

    # 1) 재산보전처분일: "보전처분결정" (신청서/송달 제외)
    보전 = pick_first(lambda c, r: "보전처분" in c and "결정" in c
                      and "신청" not in c and "에게" not in c and "송달" not in c)

    # 2) 개시결정일: "개시결정-관리인선임" 우선, 없으면 개시결정(결정) (통지/공고 제외)
    개시 = pick_first(lambda c, r: "개시결정" in c and "관리인선임" in c
                      and "에게" not in c and "송달" not in c and "공고" not in c)
    if 개시 == "미확인":
        개시 = pick_first(lambda c, r: "개시결정" in c and "결정" in c
                          and "에게" not in c and "송달" not in c and "공고" not in c
                          and "신청" not in c and "의견서" not in c and "요청" not in c)

    # 3) 시부인표 제출일: "시부인표 제출" (정정/연장/연기 제외)
    시부인 = pick_first(lambda c, r: "시부인표" in c and "제출" in c
                        and "정정" not in c and "연장" not in c and "연기" not in c
                        and "에게" not in c)

    # 4) 조사보고서 제출일: "조사보고서 제출" (연장/연기/기한/기간/신청 제외)
    조사 = pick_first(lambda c, r: "조사보고서" in c and "제출" in c
                      and "연장" not in c and "연기" not in c and "기한" not in c
                      and "기간" not in c and "에게" not in c and "신청" not in c
                      and "명령" not in c)

    # 5) 회생계획안 제출일: "회생계획안 제출" (기간/연장/요지/지정/신청 제외)
    계획 = pick_first(lambda c, r: "회생계획안" in c and "제출" in c
                      and "기간" not in c and "연장" not in c and "에게" not in c
                      and "요지" not in c and "지정" not in c and "신청" not in c)

    # 6) 인가/폐지/종결일: 법원 종국 판단만 (신청/제출/등기소 노이즈 제거)
    인가폐지, 상태 = _extract_disposition(rows)

    return {
        "재산보전처분일": 보전,
        "개시결정일":     개시,
        "시부인표 제출일": 시부인,
        "조사보고서 제출일": 조사,
        "회생계획안 제출일": 계획,
        "인가폐지일":     인가폐지,
        "인가폐지상태":   상태,
    }


def _extract_disposition(rows):
    """인가/폐지/종결/기각/각하 확정. (일자, 상태문구) 반환."""
    # 1) '종국 :' 줄 최우선 (결과가 가장 명확)
    for 일자, c, r in rows:
        if c.startswith("종국:"):
            return 일자, c.replace("종국:", "")

    # 2) 법원 종국 결정문 (당사자 제출·신청, 등기소 서류, 통지·공고 배제)
    def is_court_decision(c):
        return ("제출" not in c and "신청" not in c and "에게" not in c
                and "등기소" not in c and "등기관" not in c and "공고" not in c)

    for kw, label in [("인가결정", "인가"), ("폐지결정", "폐지"),
                      ("종결결정", "종결"), ("기각결정", "기각"), ("각하결정", "각하")]:
        for 일자, c, r in rows:
            if kw in c and is_court_decision(c):
                return 일자, label

    return "미확인", ""


# ★ v10: 총괄표 시트 생성 (파이썬 직접 추출 방식)
def build_summary_sheet(wb, cases, results):
    """맨 앞에 '회생차주 총괄표' 시트 생성.
    A:차주일련번호 B:차주명 C:법원 D:사건번호
    E:신청일  F:재산보전처분일  G:개시결정일
    H:시부인표  I:조사보고서  J:회생계획안  제출일
    K:인가일/폐지(종결)일 (+ 상태)
    ★ XLOOKUP 수식이 아니라 진행내용을 직접 판독해 확정 날짜를 기입."""
    ws = wb.create_sheet(title=SUMMARY_SHEET_NAME, index=0)

    headers = ["차주일련번호", "차주명(실명)", "관할법원", "회생사건번호",
               "신청일", "재산보전처분일", "개시결정일",
               "시부인표 제출일", "조사보고서 제출일", "회생계획안 제출일",
               "인가일 / 폐지(종결)일"]

    header_fill = PatternFill(fill_type="solid", fgColor="999999")
    header_font = Font(name="맑은 고딕", size=10, bold=True, color="FFFFFF")
    body_font   = Font(name="맑은 고딕", size=10)
    warn_font   = Font(name="맑은 고딕", size=10, color="C00000")  # 미확인 강조
    center      = Alignment(horizontal="center", vertical="center")
    thin = Side(style="thin", color="BFBFBF")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for c, h in enumerate(headers, 1):
        cell = ws.cell(1, c, h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center
        cell.border = border

    # 사건번호 → (일련번호, 차주명) 매핑 (INPUT 기준)
    meta = {}
    for cs in cases:
        meta[cs["사건번호"]] = (cs.get("일련번호", ""), cs.get("차주명", ""))

    row_idx = 2
    for r in results:
        case_str = r.get("사건번호", "")
        court    = r.get("관할법원", "")
        basic    = r.get("기본내용", {})
        progress = r.get("진행내용", [])
        serial, name = meta.get(case_str, ("", ""))

        # 신청일 = 진행내용 첫 행 일자(가장 오래된 접수), 없으면 기본내용 접수일
        신청일 = ""
        if progress:
            신청일 = str(progress[0][0]).strip() if progress[0] and progress[0][0] else ""
        if not 신청일:
            신청일 = basic.get("접수일", "")
        if not 신청일:
            신청일 = "미확인"

        dates = extract_key_dates(progress)
        인가폐지 = dates["인가폐지일"]
        if 인가폐지 not in ("", "미확인") and dates["인가폐지상태"]:
            인가폐지 = f'{인가폐지} ({dates["인가폐지상태"]})'

        values = [
            serial, name, court, case_str,
            신청일,
            dates["재산보전처분일"],
            dates["개시결정일"],
            dates["시부인표 제출일"],
            dates["조사보고서 제출일"],
            dates["회생계획안 제출일"],
            인가폐지,
        ]

        for c, v in enumerate(values, 1):
            cell = ws.cell(row_idx, c, v)
            cell.alignment = center
            cell.border = border
            cell.font = warn_font if v == "미확인" else body_font

        row_idx += 1

    # 열 너비
    widths = {"A": 11, "B": 22, "C": 14, "D": 15, "E": 13, "F": 14,
              "G": 13, "H": 14, "I": 15, "J": 15, "K": 20}
    for col, w in widths.items():
        ws.column_dimensions[col].width = w

    ws.freeze_panes = "A2"
    return ws


def save_results(results, path, cases=None):
    if not results:
        return

    # DataFrame 구성
    rows = []
    for r in results:
        basic    = r.get("기본내용", {})
        progress = r.get("진행내용", [])
        case_str = r.get("사건번호", "")
        court    = r.get("관할법원", "")

        if not progress:
            rows.append({
                "사건번호": case_str, "관할법원": court,
                "일자": "", "내용": "", "결과": "", "공시문": "",
                "내용_rgb": "rgb(0,0,0)",
                "조회결과": r.get("결과", ""), "오류내용": r.get("오류내용", ""),
                "조회일시": r.get("조회일시", ""),
                **{k: basic.get(k, "") for k in ["사건명", "신청인", "채권자", "재판부", "접수일", "종국결과", "채무자"]},
            })
        else:
            for pd_row in progress:
                rows.append({
                    "사건번호": case_str, "관할법원": court,
                    "일자":   pd_row[0] if len(pd_row) > 0 else "",
                    "내용":   pd_row[1] if len(pd_row) > 1 else "",
                    "결과":   pd_row[2] if len(pd_row) > 2 else "",
                    "공시문": pd_row[3] if len(pd_row) > 3 else "",
                    "내용_rgb": pd_row[4] if len(pd_row) > 4 else "rgb(0,0,0)",
                    "조회결과": r.get("결과", ""), "오류내용": r.get("오류내용", ""),
                    "조회일시": r.get("조회일시", ""),
                    **{k: basic.get(k, "") for k in ["사건명", "신청인", "채권자", "재판부", "접수일", "종국결과", "채무자"]},
                })

    df = pd.DataFrame(rows)
    df = compute_flags(df)

    # openpyxl 저장
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    label_fill        = PatternFill(fill_type="solid", fgColor="D9E1F2")
    label_font        = Font(bold=True, name="Arial", size=9)
    header_fill       = PatternFill(fill_type="solid", fgColor="2F5496")
    header_font_white = Font(bold=True, color="FFFFFF", name="Arial", size=9)
    center            = Alignment(horizontal="center", vertical="center")
    wrap              = Alignment(wrap_text=True, vertical="top")

    HEADER_ROW_IDX = 8  # 기본내용 4행 + 공백(5행 조회메타 포함) + 헤더 8행

    for r in results:
        case_str = r.get("사건번호", "")
        court    = r.get("관할법원", "")
        basic    = r.get("기본내용", {})

        sheet_name = make_sheet_name(court, case_str)
        ws = wb.create_sheet(title=sheet_name)

        # 기본내용 (1~4행)
        rows_basic = [
            [("사건번호", basic.get("사건번호", case_str)), ("사건명",   basic.get("사건명", ""))],
            [("신청인",   basic.get("신청인", basic.get("채무자", ""))), ("채권자",   basic.get("채권자", ""))],
            [("재판부",   basic.get("재판부", "")),                       ("관할법원", court)],
            [("접수일",   basic.get("접수일", "")),                       ("종국결과", basic.get("종국결과", ""))],
        ]
        for ri, pairs in enumerate(rows_basic, 1):
            for ci_pair, (label, value) in enumerate(pairs):
                ci = ci_pair * 2 + 1
                cl = ws.cell(ri, ci, label)
                cl.font = label_font
                cl.fill = label_fill
                cl.alignment = center
                cv = ws.cell(ri, ci + 1, value)
                cv.font = Font(name="Arial", size=9)

        # 조회일시/결과 (5행)
        for ci_pair, (label, value) in enumerate([
            ("조회일시", r.get("조회일시", "")),
            ("조회결과", r.get("결과", ""))
        ]):
            ci = ci_pair * 2 + 1
            cl = ws.cell(5, ci, label)
            cl.font = label_font
            cl.fill = label_fill
            cl.alignment = center
            cv = ws.cell(5, ci + 1, value)
            cv.font = Font(name="Arial", size=9)

        # 진행내용 헤더 (8행)
        all_headers = ["일자", "내용", "결과", "공시문"] + FLAG_HEADER_COLS
        for ci, h in enumerate(all_headers, 1):
            cell = ws.cell(HEADER_ROW_IDX, ci, h)
            cell.font = header_font_white
            cell.fill = header_fill
            cell.alignment = center

        # 해당 사건 데이터
        case_df = df[df["사건번호"] == case_str].reset_index(drop=True)

        for ri, row in case_df.iterrows():
            excel_row = HEADER_ROW_IDX + 1 + ri
            rgb_str   = row.get("내용_rgb", "rgb(0,0,0)")
            hex_color = rgb_to_hex(rgb_str)

            # 일자/내용/결과/공시문 (색상 적용)
            for ci, col in enumerate(["일자", "내용", "결과", "공시문"], 1):
                c = ws.cell(excel_row, ci, row.get(col, ""))
                c.alignment = wrap
                c.font = Font(name="Arial", size=9, color=hex_color)

            # 플래그 컬럼
            for ci, col in enumerate(FLAG_HEADER_COLS, 5):
                c = ws.cell(excel_row, ci, row.get(col, ""))
                c.alignment = center
                c.font = Font(name="Arial", size=9)

        # 열 너비
        ws.column_dimensions["A"].width = 14
        ws.column_dimensions["B"].width = 55
        ws.column_dimensions["C"].width = 22
        ws.column_dimensions["D"].width = 12
        for ci in range(5, len(all_headers) + 1):
            col_letter = openpyxl.utils.get_column_letter(ci)
            ws.column_dimensions[col_letter].width = 10
        # G열(7)~S열(19) 숨김 처리
        for ci in range(7, 20):
            col_letter = openpyxl.utils.get_column_letter(ci)
            ws.column_dimensions[col_letter].hidden = True

        # 8행 기준 자동필터
        ws.auto_filter.ref = f"A{HEADER_ROW_IDX}:{openpyxl.utils.get_column_letter(len(all_headers))}{HEADER_ROW_IDX}"

    # ★ v10: 맨 앞에 총괄표 시트 생성 (개별 시트 다 만든 뒤에 삽입해야 참조 유효)
    if cases is not None:
        build_summary_sheet(wb, cases, results)

    wb.save(path)
    print(f"저장 완료: {path}")


# ============================================================
# 메인
# ============================================================
def run_crawl(cases, api_key, btpr_nm, output_path,
              on_log=None, on_progress=None, on_case_done=None,
              should_stop=None, delay=None):
    """크롤링 코어. GUI/CLI 공용.

    파라미터
      cases       : [{"일련번호","차주명","법원","사건번호"}, ...]
      api_key     : 보안문자 해제 API 키
      btpr_nm     : 당사자명
      output_path : 결과 저장 경로(.xlsx)
      on_log(msg)            : 로그 한 줄 콜백
      on_progress(done,total): 진행률 콜백 (사건 시작 시)
      on_case_done(idx,total,result,elapsed): 사건 1건 완료 콜백
      should_stop()          : True 반환 시 루프 중단
      delay                  : 사건 간 대기(기본 DELAY_BETWEEN_REQUESTS)

    반환: (results, summary_dict)
    """
    def log(msg):
        if on_log:
            on_log(msg)
        else:
            print(msg)

    if delay is None:
        delay = DELAY_BETWEEN_REQUESTS

    total = len(cases)
    log("=" * 55)
    log("  대법원 회생사건 자동 조회 v10")
    log("=" * 55)

    if not api_key:
        log("❌ 인증 API 키가 없습니다.")
        return [], {"ok": 0, "none": 0, "err": 0, "total": total, "path": None}

    log(f"✅ API 키 확인 완료 ({api_key[:6]}...)")
    log(f"📋 조회 대상: 총 {total}건 / 당사자명: '{btpr_nm}'")
    log("🌐 브라우저 실행 중...")

    driver = setup_driver()
    results = []

    try:
        for i, case in enumerate(cases, 1):
            if should_stop and should_stop():
                log("\n⏹ 사용자 중단")
                break

            if on_progress:
                on_progress(i - 1, total)

            log(f"\n[{i}/{total}] {case['법원']} / {case['사건번호']}")
            t0 = time.time()
            r = search_one(driver, api_key, case["법원"], case["사건번호"], btpr_nm=btpr_nm)
            elapsed = time.time() - t0
            results.append(r)

            prog_count = len(r.get("진행내용", []))
            결과 = r.get("결과", "")
            오류 = r.get("오류내용", "")
            log(f"  → {결과} {오류} | 진행내용 {prog_count}건 | {elapsed:.1f}초")

            if on_case_done:
                on_case_done(i, total, r, elapsed)
            if on_progress:
                on_progress(i, total)

            if i % 5 == 0:
                save_results(results, output_path, cases=cases)
                log(f"  💾 중간 저장 ({i}건)")

            time.sleep(delay)

    except KeyboardInterrupt:
        log("\n⏹ 중단됨 (Ctrl+C)")
    finally:
        try:
            driver.quit()
        except Exception:
            pass

    save_results(results, output_path, cases=cases)

    ok   = sum(1 for r in results if r.get("결과") == "조회성공")
    none = sum(1 for r in results if r.get("결과") == "사건 없음")
    err  = len(results) - ok - none
    log(f"\n완료 — 성공: {ok}건 / 사건없음: {none}건 / 오류: {err}건")
    log(f"결과파일: {output_path}")

    return results, {"ok": ok, "none": none, "err": err, "total": total, "path": output_path}


def main():
    """CLI 실행 (엑셀 파일 경로 기반, 기존 방식 유지)"""
    if not VERIFY_API_KEY:
        print("❌ .env_verify 파일에 VERIFY_API_KEY가 없습니다.")
        return
    print(f"\n엑셀 로드: {INPUT_EXCEL_PATH}")
    try:
        cases = load_cases(INPUT_EXCEL_PATH, INPUT_SHEET_NAME,
                           COL_SERIAL, COL_NAME, COL_COURT, COL_CASE_NUMBER, HEADER_ROW)
    except FileNotFoundError:
        print(f"❌ 파일 없음: {INPUT_EXCEL_PATH}")
        return
    run_crawl(cases, VERIFY_API_KEY, BTPR_NM, OUTPUT_EXCEL_PATH)


if __name__ == "__main__":
    main()
